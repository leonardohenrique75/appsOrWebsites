# -*- coding: utf-8 -*-
"""
Created on Thu Jul 23 17:30:12 2026

@author: Leo
"""
from pathlib import Path
from openpyxl import load_workbook

#troubleshooting
import os
print(os.path.abspath("games.json"))

import json #to read my personal database
with open("games.json", "r", encoding="utf-8") as file:
    games = json.load(file)
        
if isinstance(games, dict):
    games = [games]
    
#function below is to read the excel, using openpyxl and pathlib
def import_excel_to_json():
    excel_file = Path("games.xlsx")
#error prevention
    if not excel_file.exists():
        print("Archive games.xlsx not found.")
        return

    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    # Using Excel headers to match the JSON headers: name, system, genre, status, grade
    headers = [cell.value for cell in sheet[1]]

    # Not repeating JSON Games, not repeating existing games is the functionality of the code below
    existing_games = {
        (game["name"].strip().lower(), game["system"].strip().lower())
        for game in games
    }

    added = 0

    # Starting in line two, where the gamaes begins
    for row in sheet.iter_rows(min_row=2, values_only=True):
        game = dict(zip(headers, row))

        # Ignoring the reminescent lines that are empty
        if not game.get("name"):
            continue

        game["name"] = str(game["name"]).strip()
        game["system"] = str(game.get("system") or "").strip()
        game["genre"] = str(game.get("genre") or "").strip()
        game["status"] = str(game.get("status") or "").strip()
        game["grade"] = float(game.get("grade") or 0)

        game_id = (game["name"].lower(), game["system"].lower())
        #if the game does not exist on the list/dict/JSON, add it
        if game_id not in existing_games:
            games.append(game)
            existing_games.add(game_id)
            added += 1
    #calling save_games() to update the list with the new value/game
    save_games()
    print(f"{added} Game(s) imported from Excel.")
#use the function below in hashtag to import the games from an Excel to the games.json   
#import_excel_to_json()
#function to save games when called by add_game function, no params
def save_games():
    with open("games.json", "w", encoding="utf-8") as file:
        json.dump(games, file, ensure_ascii=False, indent=2)
        
import unicodedata

#function that solves the duplicate problem
def fix_catalog():
    cleaned_games = []
    existing_games = set()

    for game in games:
        # Drags the sytem typo column to system
        if not game.get("system") and game.get("sytem"):
            game["system"] = game["sytem"]

        game.pop("sytem", None)

        # Solving accentuations problmens
        game["name"] = game["name"].replace("PokÃ©mon", "Pokémon")

        game_id = (
            unicodedata.normalize("NFKD", game["name"])
            .encode("ASCII", "ignore")
            .decode("ASCII")
            .lower(),
            game["system"].lower()
        )

        # mantains only the first occourence of a game
        if game_id not in existing_games:
            cleaned_games.append(game)
            existing_games.add(game_id)

    games.clear()
    games.extend(cleaned_games)
    save_games()

fix_catalog()
#function that adds a new game to the catalog 
def add_game():
    new_game = {
        "name": input("Enter Game name: "),
        "system": input("Enter Gaming System: "),
        "genre": input("Enter Game genre: "),
        "status": input("Status: "),
        "grade": float(input("Set a grade from 0 to 10: "))
}
#appends a new game to the JSON
    games.append(new_game)
#calls save_games to update the database
    save_games()
#message that shows on the screen after saving a game
    print("Game entry saved successfully!")


#with open("games.json", "w", encoding="utf-8") as file:
#    json.dump(games, file, ensure_ascii=False, indent=2)
    
def list_games():
#if there is no games on the JSON, do the following
    if not games:
        print("No game entry.")
        return
        
    print("\n --- My games ---")
#shows games with the registry entry and the informations mentioned
    for number, game in enumerate(games, start=1):
        print(f"\n{number}. {game['name']}")
        print(f"Gaming system: {game['system']}")
        print(f"Game genre: {game['genre']}")
        print(f"Grade: {game['grade']}")
    
#a while that gives the user 3 options, add, list or exit
while True:
    print("\n--- Game Catalog ---")
    print("1 - Add game")
    print("2 - Game list")
    print("3 - Exit")
    
    option = input("Choose an option: ")
    
    if option == "1":
        add_game()
    elif option == "2":
        list_games()
    elif option == "3":
        print("See ya, Pokémon trainer!")
    #break here ends the while
        break
    #error-proof system
    else:
        print("Error. Choose 1, 2 or 3.")